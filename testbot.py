import praw
import sqlite3
import os
from openpyxl import Workbook, load_workbook

conn = sqlite3.connect('bot.db')
c = conn.cursor()

reddit = praw.Reddit("bot1", user_agent="Comment Extraction (by /u/username)")

subreddits = ["python"]

limit = 3

count = 0

filename = "redditors.xlsx"

subject = "TEST"
message = "test message"


def connectDB():
    c.execute("""CREATE TABLE if not exists users (username text)""")
    print("Connected DB successfully")


def is_new_user(author):
    c.execute("SELECT * FROM users WHERE username=:author",
              {"author": author})
    return c.fetchone() == None


def insert_user(author):
    c.execute("INSERT INTO users VALUES (:author)", {"author": author})
    conn.commit()
    global count
    count += 1


def add_to_sheet(sheet, author, user_type):
    sheet.insert_rows(idx=2)
    sheet["A2"] = '=HYPERLINK("{}", "{}")'.format(
        "https://www.reddit.com/user/" + author, author)
    sheet["A2"].style = "Hyperlink"
    sheet["B2"] = user_type


def process_comments(objects, workbook, subreddit):
    for object in objects:
        if type(object).__name__ == "Comment":
            process_comments(object.replies, workbook, subreddit)
            author = str(object.author)
            if is_new_user(author):
                print("Found new user {}".format(author))
                # count += 1

                insert_user(author)

                if not subreddit in workbook.sheetnames:
                    sheet = workbook.create_sheet(subreddit)
                    sheet["A1"] = "username"
                    sheet["B1"] = "type"

            sheet = workbook[subreddit]
            add_to_sheet(sheet, author, "Commenter")
            print("Sending message to...{}".format(author))
            reddit.redditor(author).message(subject, message)
            workbook.save(filename=filename)

            print("Saved {} to {}".format(author, filename))

        elif type(object).__name__ == "MoreComments":
            process_comments(object.comments(), workbook, subreddit)


def fetch_new_users(workbook):
    print("Fetching... new users...")
    i = 1
    for post_id in reddit.subreddit("+".join(subreddits)).new(limit=limit):
        print("Fetching... {} of {}".format(i, limit))
        i += 1

        submission = reddit.submission(id=post_id)
        author = str(submission.author)
        subreddit = str(submission.subreddit)

        print("Checking post comments sections...")
        process_comments(submission.comments.list(), workbook, subreddit)

        if is_new_user(author):
            print("Found new user {}".format(author))
            # count += 1

            insert_user(author)

            if not subreddit in workbook.sheetnames:
                sheet = workbook.create_sheet(subreddit)
                sheet["A1"] = "username"
                sheet["B1"] = "type"
            sheet = workbook[subreddit]
            add_to_sheet(sheet, author, "Original Poster")

            print("Sending message to...{}".format(author))
            reddit.redditor(author).message(subject, message)

            workbook.save(filename=filename)

            print("Saved {} to {}".format(author, filename))

    return count


def load_xlxs(filename):
    if os.path.exists(filename):
        return load_workbook(filename=filename)
    return Workbook()


def main():
    connectDB()

    workbook = load_xlxs(filename)

    fetch_new_users(workbook)

    print("Found {} new users. Added to redditors.xlsx".format(count))

    # fetch_new_commentors(reddit.submission(id="ie7fje"))

    conn.commit()
    conn.close()


if __name__ == '__main__':
    main()
